import openpyxl
import random


def get_account_by_random():
    accounts = get_accounts()
    count = random.randint(0, len(accounts)-1)
    print u'use account:'
    print accounts[count]
    return accounts[count]


def get_accounts():
    accounts_list = list()
    # weibo accounts instead zhihu accounts
    work_book = openpyxl.load_workbook(r'accounts.xlsx')
    work_sheet = work_book.get_sheet_by_name('Sheet1')
    # print work_sheet.get_highest_column()
    for r in xrange(0, work_sheet.get_highest_row() + 1):
        key = work_sheet.cell(row=r, column=1).value
        word = work_sheet.cell(row=r, column=2).value
        if key != None and word != None:
            # print key, word
            accounts_list.append((key, word))
    return accounts_list[1:]

if __name__ == '__main__':
    accounts = get_accounts()
    for account in accounts:
        print account